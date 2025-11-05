import openpyxl

wb = openpyxl.load_workbook('quest.xlsx')
ws = wb.active

print("Excel structure with groups:\n")
print("="*100)

for i in range(1, min(30, ws.max_row + 1)):
    col1 = ws.cell(i, 1).value
    col2 = ws.cell(i, 2).value
    col3 = ws.cell(i, 3).value
    
    if col1:
        print(f"\n[GROUP] Row {i}, Col 1: {col1}")
    if col2:
        print(f"  [QUESTION] Row {i}, Col 2: {str(col2)[:80]}...")
    if col3:
        print(f"  [SUBQ] Row {i}, Col 3: {str(col3)[:80]}...")

wb.close()
