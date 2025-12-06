import sqlite3
import re
import openpyxl
from typing import Dict, List, Tuple
from bs4 import BeautifulSoup

def parse_excel_questions_and_groups(excel_path: str) -> Tuple[Dict[str, str], Dict[str, Dict]]:
    """Parse quest.xlsx and extract questions and group information.
    
    Returns:
        Tuple of (questions_dict, groups_dict)
        - questions_dict: Maps question numbers to their full question text
        - groups_dict: Maps group numbers to {'name': str, 'question_numbers': list}
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    questions = {}
    groups = {}
    current_group_num = None
    
    for row in range(1, ws.max_row + 1):
        # Check column 1 (group headers)
        group_cell = ws.cell(row, 1).value
        if group_cell and isinstance(group_cell, str):
            # Extract group number and name (e.g., "1. Границы клауз...")
            match = re.match(r'^(\d+)\.\s*(.+)$', group_cell.strip())
            if match:
                group_num = match.group(1)
                group_name = match.group(2).strip()
                current_group_num = group_num
                groups[group_num] = {
                    'name': group_name,
                    'question_numbers': []
                }
        
        # Check column 2 (main questions)
        question_cell = ws.cell(row, 2).value
        if question_cell and isinstance(question_cell, str):
            question_text = question_cell.strip()
            match = re.match(r'^(\d+\.\d+(?:\.\d+)*)[.\s]', question_text)
            if match:
                question_num = match.group(1)
                questions[question_num] = question_text
                # Associate question with current group
                if current_group_num and current_group_num in groups:
                    groups[current_group_num]['question_numbers'].append(question_num)
        
        # Check column 3 (subquestions)
        subquestion_cell = ws.cell(row, 3).value
        if subquestion_cell and isinstance(subquestion_cell, str):
            subquestion_text = subquestion_cell.strip()
            match = re.match(r'^(\d+\.\d+(?:\.\d+)*)[.\s]', subquestion_text)
            if match:
                subquestion_num = match.group(1)
                questions[subquestion_num] = subquestion_text
                # Associate subquestion with current group
                if current_group_num and current_group_num in groups:
                    groups[current_group_num]['question_numbers'].append(subquestion_num)
    
    wb.close()
    return questions, groups

def parse_excel_questions(excel_path: str) -> Dict[str, str]:
    """Parse quest.xlsx and extract all questions with their numbers and full text.
    
    Returns a dictionary mapping question numbers to their full question text.
    """
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    
    questions = {}
    
    for row in range(1, ws.max_row + 1):
        # Check column 2 (main questions)
        question_cell = ws.cell(row, 2).value
        if question_cell and isinstance(question_cell, str):
            question_text = question_cell.strip()
            # Extract question number (e.g., "1.1.", "2.4.")
            match = re.match(r'^(\d+\.\d+(?:\.\d+)*)[.\s]', question_text)
            if match:
                question_num = match.group(1)
                questions[question_num] = question_text
        
        # Check column 3 (subquestions)
        subquestion_cell = ws.cell(row, 3).value
        if subquestion_cell and isinstance(subquestion_cell, str):
            subquestion_text = subquestion_cell.strip()
            # Extract subquestion number (e.g., "2.4.1.", "2.12.")
            match = re.match(r'^(\d+\.\d+(?:\.\d+)*)[.\s]', subquestion_text)
            if match:
                subquestion_num = match.group(1)
                questions[subquestion_num] = subquestion_text
    
    wb.close()
    return questions

def normalize_question_text(text: str) -> str:
    """Normalize question text for matching: remove extra spaces, normalize whitespace."""
    # Replace multiple spaces/tabs with single space
    text = re.sub(r'[\s]+', ' ', text)
    return text.strip()

def parse_language_file(filename: str, questions: Dict[str, str]) -> Dict[str, str]:
    """Parse a language HTML file with <answer></answer> tags containing HTML content.
    
    Args:
        filename: Path to the language HTML file in languages/ folder
        questions: Dictionary of question numbers to question texts from Excel
    
    Returns:
        Dictionary mapping question numbers to HTML answer texts (unchanged from file)
    """
    # Try different encodings
    encodings = ['utf-8', 'utf-16', 'cp1251', 'latin-1']
    content = None
    
    for encoding in encodings:
        try:
            with open(filename, 'r', encoding=encoding) as f:
                content = f.read()
            break
        except (UnicodeDecodeError, UnicodeError):
            continue
    
    if content is None:
        raise ValueError(f"Could not decode file {filename} with any known encoding")
    
    answers = {}
    
    # Find all <answer>...</answer> blocks
    # Pattern: question_number followed by <answer>content</answer>
    pattern = r'(\d+(?:\.\d+)*)[.\s\t]*<answer>(.*?)</answer>'
    
    for match in re.finditer(pattern, content, re.DOTALL | re.IGNORECASE):
        question_num = match.group(1)
        answer_html = match.group(2).strip()
        
        # Only process if this question number is in our questions dict
        if question_num in questions:
            # Store the HTML content as-is (no conversion needed)
            answers[question_num] = answer_html
    
    return answers

def convert_answer_to_html(text: str) -> str:
    """Convert plain text answer to HTML with proper formatting.
    
    Handles:
    - Bullet points (•) -> <ul><li>
    - Multiple paragraphs -> <p>
    - Preserves special characters and formatting
    """
    if not text:
        return ''
    
    lines = text.split('\n')
    html_parts = []
    in_list = False
    current_paragraph = []
    
    for line in lines:
        line = line.strip()
        if not line:
            # Empty line - end current paragraph or list
            if current_paragraph:
                if in_list:
                    html_parts.append('</ul>')
                    in_list = False
                else:
                    html_parts.append(f'<p>{" ".join(current_paragraph)}</p>')
                current_paragraph = []
            continue
        
        # Check if line starts with bullet
        if line.startswith('•'):
            # Start list if not already in one
            if not in_list:
                if current_paragraph:
                    html_parts.append(f'<p>{" ".join(current_paragraph)}</p>')
                    current_paragraph = []
                html_parts.append('<ul>')
                in_list = True
            
            # Add list item
            item_text = line[1:].strip()
            html_parts.append(f'<li>{item_text}</li>')
        else:
            # Regular text
            if in_list:
                # End the list
                html_parts.append('</ul>')
                in_list = False
            current_paragraph.append(line)
    
    # Handle remaining content
    if in_list:
        html_parts.append('</ul>')
    elif current_paragraph:
        html_parts.append(f'<p>{" ".join(current_paragraph)}</p>')
    
    return ''.join(html_parts)

def merge_subquestions(language_data: Dict[str, str], quest_data: Dict[str, str]) -> Dict[str, str]:
    """Merge sub-questions that don't exist in quest.txt into their parent questions.
    
    For example, if 5.3.1, 5.3.2 exist in language file but not in quest.txt,
    merge them into 5.3 as: "5.3 text. 5.3.1: sub text. 5.3.2: sub text."
    """
    merged = {}
    quest_numbers = set(quest_data.keys())
    
    for num, text in language_data.items():
        if num in quest_numbers:
            # This question exists in quest.txt, keep it as is
            merged[num] = text
        else:
            # This is a sub-question not in quest.txt, find parent
            parts = num.split('.')
            # Try to find parent by removing last component
            while len(parts) > 1:
                parts = parts[:-1]
                parent_num = '.'.join(parts)
                if parent_num in quest_numbers:
                    # Append to parent question
                    if parent_num not in merged:
                        merged[parent_num] = language_data.get(parent_num, '')
                    # Add sub-question as continuation
                    if merged[parent_num]:
                        merged[parent_num] += f" {num}: {text}"
                    else:
                        merged[parent_num] = f"{num}: {text}"
                    break
    
    return merged

def get_group_number(question_number: str) -> str:
    """Extract the top-level group number from a question number."""
    return question_number.split('.')[0]

def create_database(db_path: str):
    """Create SQLite database with questionnaire data."""
    
    print("Parsing Excel file for questions and groups...")
    quest_data, groups_data = parse_excel_questions_and_groups('quest.xlsx')
    print(f"Found {len(quest_data)} questions in quest.xlsx")
    print(f"Found {len(groups_data)} groups in quest.xlsx")
    
    print("\nParsing language files...")
    russian_data = parse_language_file('languages/01-russian.html', quest_data)
    print(f"Found {len(russian_data)} answers in 01-russian.html")
    
    muira_data = parse_language_file('languages/02-muira.html', quest_data)
    print(f"Found {len(muira_data)} answers in 02-muira.html")
    
    danish_data = parse_language_file('languages/03-danish.html', quest_data)
    print(f"Found {len(danish_data)} answers in 03-danish.html")
    
    nganasan_data = parse_language_file('languages/04-nganasan.html', quest_data)
    print(f"Found {len(nganasan_data)} answers in 04-nganasan.html")
    
    westcircassian_data = parse_language_file('languages/05-westcircassian.html', quest_data)
    print(f"Found {len(westcircassian_data)} answers in 05-westcircassian.html")
    
    polish_data = parse_language_file('languages/06-polish.html', quest_data)
    print(f"Found {len(polish_data)} answers in 06-polish.html")
    
    bulgarian_data = parse_language_file('languages/07-bulgarian.html', quest_data)
    print(f"Found {len(bulgarian_data)} answers in 07-bulgarian.html")
    
    nanai_data = parse_language_file('languages/08-nanai.html', quest_data)
    print(f"Found {len(nanai_data)} answers in 08-nanai.html")
    
    nornakhichevan_data = parse_language_file('languages/09-nor-nakhichevan.html', quest_data)
    print(f"Found {len(nornakhichevan_data)} answers in 09-nor-nakhichevan.html")
    
    udmurt_data = parse_language_file('languages/10-udmurt.html', quest_data)
    print(f"Found {len(udmurt_data)} answers in 10-udmurt.html")
    
    greben_data = parse_language_file('languages/11-Greben.html', quest_data)
    print(f"Found {len(greben_data)} answers in 11-Greben.html")
    
    mountmari_data = parse_language_file('languages/12-mountmari.html', quest_data)
    print(f"Found {len(mountmari_data)} answers in 12-mountmari.html")
    
    icari_data = parse_language_file('languages/13-icari.html', quest_data)
    print(f"Found {len(icari_data)} answers in 13-icari.html")
    
    macedonian_data = parse_language_file('languages/14-macedonian.html', quest_data)
    print(f"Found {len(macedonian_data)} answers in 14-macedonian.html")
    
    norwegian_data = parse_language_file('languages/15-norwegian.html', quest_data)
    print(f"Found {len(norwegian_data)} answers in 15-norwegian.html")
    
    kumyk_data = parse_language_file('languages/16-kumyk.html', quest_data)
    print(f"Found {len(kumyk_data)} answers in 16-kumyk.html")
    
    northernkhanty_data = parse_language_file('languages/17-northernkhanty.html', quest_data)
    print(f"Found {len(northernkhanty_data)} answers in 17-northernkhanty.html")
    
    ulch_data = parse_language_file('languages/18-ulch.html', quest_data)
    print(f"Found {len(ulch_data)} answers in 18-ulch.html")
    
    # Create database
    conn = sqlite3.connect(db_path)
    cursor = conn.cursor()
    
    # Create groups table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS groups (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            group_number TEXT UNIQUE NOT NULL,
            group_name TEXT NOT NULL
        )
    ''')
    
    # Create questions table with TEXT columns for HTML content
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS questions (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            question_number TEXT UNIQUE NOT NULL,
            group_id INTEGER NOT NULL,
            question_text TEXT NOT NULL,
            russian TEXT,
            muira TEXT,
            danish TEXT,
            nganasan TEXT,
            westcircassian TEXT,
            polish TEXT,
            bulgarian TEXT,
            nanai TEXT,
            nornakhichevan TEXT,
            udmurt TEXT,
            greben TEXT,
            mountmari TEXT,
            icari TEXT,
            macedonian TEXT,
            norwegian TEXT,
            kumyk TEXT,
            northernkhanty TEXT,
            ulch TEXT,
            FOREIGN KEY (group_id) REFERENCES groups(id)
        )
    ''')
    
    # Create indexes for faster queries
    cursor.execute('''
        CREATE INDEX IF NOT EXISTS idx_group_id ON questions(group_id)
    ''')
    
    # Insert groups
    print("\nInserting groups into database...")
    group_id_map = {}  # Map group_number to database id
    for group_num in sorted(groups_data.keys(), key=lambda x: int(x)):
        group_info = groups_data[group_num]
        cursor.execute('''
            INSERT INTO groups (group_number, group_name)
            VALUES (?, ?)
        ''', (group_num, group_info['name']))
        group_id_map[group_num] = cursor.lastrowid
        print(f"  Group {group_num}: {group_info['name']}")
    
    # Insert questions
    print("\nInserting questions into database...")
    for question_num, question_text in quest_data.items():
        group_num = get_group_number(question_num)
        
        # Get group_id from the map
        group_id = group_id_map.get(group_num)
        if not group_id:
            print(f"  Warning: No group found for question {question_num}")
            continue
        
        # Get language data (use empty string if not found)
        russian_text = russian_data.get(question_num, '')
        muira_text = muira_data.get(question_num, '')
        danish_text = danish_data.get(question_num, '')
        nganasan_text = nganasan_data.get(question_num, '')
        westcircassian_text = westcircassian_data.get(question_num, '')
        polish_text = polish_data.get(question_num, '')
        bulgarian_text = bulgarian_data.get(question_num, '')
        nanai_text = nanai_data.get(question_num, '')
        nornakhichevan_text = nornakhichevan_data.get(question_num, '')
        udmurt_text = udmurt_data.get(question_num, '')
        greben_text = greben_data.get(question_num, '')
        mountmari_text = mountmari_data.get(question_num, '')
        icari_text = icari_data.get(question_num, '')
        macedonian_text = macedonian_data.get(question_num, '')
        norwegian_text = norwegian_data.get(question_num, '')
        kumyk_text = kumyk_data.get(question_num, '')
        northernkhanty_text = northernkhanty_data.get(question_num, '')
        ulch_text = ulch_data.get(question_num, '')
        
        try:
            cursor.execute('''
                INSERT INTO questions 
                (question_number, group_id, question_text, russian, muira, danish, nganasan, westcircassian, polish, bulgarian, nanai, nornakhichevan, udmurt, greben, mountmari, icari, macedonian, norwegian, kumyk, northernkhanty, ulch)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (question_num, group_id, question_text, russian_text, muira_text, danish_text, nganasan_text, westcircassian_text, polish_text, bulgarian_text, nanai_text, nornakhichevan_text, udmurt_text, greben_text, mountmari_text, icari_text, macedonian_text, norwegian_text, kumyk_text, northernkhanty_text, ulch_text))
        except sqlite3.IntegrityError as e:
            print(f"ERROR: Duplicate question number: {question_num}")
            print(f"  Text: {question_text[:100]}")
            raise
    
    conn.commit()
    
    # Print statistics
    cursor.execute('SELECT COUNT(*) FROM questions')
    total_count = cursor.fetchone()[0]
    
    cursor.execute('SELECT COUNT(*) FROM groups')
    group_count = cursor.fetchone()[0]
    
    print(f"\nDatabase created successfully!")
    print(f"Total questions: {total_count}")
    print(f"Total groups: {group_count}")
    
    # Show some examples
    print("\nExample queries:")
    print("\n1. First 3 questions:")
    cursor.execute('SELECT question_number, question_text FROM questions LIMIT 3')
    for row in cursor.fetchall():
        print(f"  {row[0]}: {row[1][:80]}...")
    
    print("\n2. Questions in group 1:")
    cursor.execute('''
        SELECT COUNT(*) FROM questions q
        JOIN groups g ON q.group_id = g.id
        WHERE g.group_number = "1"
    ''')
    print(f"  Count: {cursor.fetchone()[0]}")
    
    print("\n3. Sample question with HTML answer:")
    cursor.execute('SELECT question_number, question_text, russian FROM questions WHERE russian != "" LIMIT 1')
    row = cursor.fetchone()
    if row:
        print(f"  {row[0]}: {row[1][:80]}...")
        print(f"  Russian answer (HTML): {row[2][:100]}...")
    
    conn.close()

if __name__ == '__main__':
    create_database('grant_database.db')
    print("\nDatabase file created: grant_database.db")
