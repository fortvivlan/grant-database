import openpyxl

# Load the Excel file
wb = openpyxl.load_workbook('quest.xlsx')
ws = wb.active

print("Excel file structure:")
print("="*100)

# Read first 20 rows to understand structure
for i in range(1, min(21, ws.max_row + 1)):
    col1 = ws.cell(i, 1).value or ""
    col2 = ws.cell(i, 2).value or ""
    col3 = ws.cell(i, 3).value or ""
    
    if col1 or col2 or col3:
        print(f"\nRow {i}:")
        if col1:
            print(f"  Column 1 (Header): {str(col1)[:100]}")
        if col2:
            print(f"  Column 2 (Question): {str(col2)[:100]}")
        if col3:
            print(f"  Column 3 (Subquestion): {str(col3)[:100]}")

print(f"\n\nTotal rows in Excel: {ws.max_row}")
wb.close()
